#Libreria Selenium
import selenium
#Libreria Webdriver
from selenium import webdriver

#Librerias Webdrivers Services de los navegadores
from selenium.webdriver.chrome.service import Service as ChromeService 
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
#Librerias Webdrivers options de los navegadores
from selenium.webdriver.chrome.options import Options as OpcionesChrome
from selenium.webdriver.firefox.options import Options as OpcionesFirefox
from selenium.webdriver.edge.options import Options as OpcionesEdge

#Librerias Webdrivers Manager de los navegadores
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager

#Librerias Webdrivers funcionalidades
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.remote.webelement import isDisplayed_js
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.alert import Alert

import cv2
import numpy as np
from mss import MSS
import threading

#from selenium.webdriver.common.print_page_options import PrintOptions as PrintOptions
#import base64
#import aspose.pdf as ap
#import urllib
from src.Function.Inicializar import Inicializar
from src.Function.DriverFactory import DriverFactory
from selenium.common.exceptions import NoSuchElementException,NoAlertPresentException,NoSuchWindowException,TimeoutException, UnexpectedAlertPresentException, WebDriverException
import json
import pytest
import time
import openpyxl
import re # para expresiones regulares
import os # para capturas
import allure
import pyodbc
from allure_commons.types import AttachmentType
from PIL import Image #Pillow - Manejo de Imagenes
from io import BytesIO #Para conocer tamaños en bytes, ya esta instalado en Python
from unittest.case import skip
from threading import Thread,Barrier

class Functions(Inicializar):
    
    Nav_utilizado_capturas = ""
    # Diccionario para mapear las estrategias de búsqueda a los valores de By
    BY = {
        "ID": By.ID,
        "NAME": By.NAME,
        "XPATH": By.XPATH,
        "CSS": By.CSS_SELECTOR,
        "CLASS": By.CLASS_NAME,
        "LINK": By.LINK_TEXT,
        "TAG": By.TAG_NAME
    }

    # Método para abrir el navegador según la configuración de Inicializar, el parámetro navegador y la opción de Selenium Grid seleccionada. Se puede abrir un navegador local, en docker o selenium grid local según la configuración.
    def abrir_navegador(self,navegador=Inicializar.Navegador, URL_SeleniumGrid = Inicializar.URL_SeleniumGrid,PortSelGrid=Inicializar.PortSelGrid):
        print(u"Directorio Base:" + Inicializar.BaseDir)
        print("-------------------------------------------")
        print(navegador)
        print("-------------------------------------------")
        self.Nav_utilizado_capturas = navegador   

        # Crear una instancia de DriverFactory con la URL y el puerto de Selenium Grid
        self.DriverFactory = DriverFactory(navegador)
        # Obtener el driver del navegador especificado
        self.driver = self.DriverFactory.get_driver()
        print(f"Se abrió el navegador {navegador} correctamente.")
        return self.driver  
    
    #Cerrar la instancia del navegador
    def cerrar_driver_navegador(self):
        if self.driver:
            self.driver.quit()
        else:
            print("No hay un driver activo para cerrar.")

    #Dirigir a la URL del sitio de pruebas  
    def get_url_driver(self,URL):
        return self.driver.get(URL)

    #Obtener Archivo JSON con los localizadores por medio del nombre      
    def obtener_archivo_json(self,file):
        json_ruta = Inicializar.Json + "/"+file+'.json'
        try:
            with open(json_ruta,'r')as read_file:
                self.json_strings = json.loads(read_file.read())
                print(u"Obtener Archivo Json: " + json_ruta)
                print(self.json_strings)
                return self.json_strings
            
        except FileNotFoundError:
            self.json_strings =False
            pytest.skip(u'Obtener Archivo Json: No se encontro el archivo json' + file)
            Functions.cerrar_driver_navegador(self)

    #Metodo para encontrar elementos en el DOM
    def Find_Element_On_DOM(self,estrategia_busqueda, valor_busqueda):
        try:
            elemento = self.driver.find_element(Functions.BY[estrategia_busqueda.upper()], valor_busqueda)
            print(u'Find Element: Se esta interactuando con el elemento ' + valor_busqueda)
            return elemento
        except TimeoutException:
            print(u'El elemento esperado no se encontro: ' + valor_busqueda)
            Functions.cerrar_driver_navegador(self)
        except NoSuchElementException:
            print(u'El elemento esperado no se encontro: ' + valor_busqueda)
            Functions.cerrar_driver_navegador(self)
     
    #Obtener entidad de elemento en el archivo JSON       
    def Get_Entity(self,page, elemento):
        try:
            if self.json_strings is False:
                print(u'Define el archivo JSON de la prueba')
            else:
                entidad = self.json_strings[page][elemento]
                print(f"Se encontraron los valores en el JSON para la entidad utilizada: " + str(elemento) + " con los valores " + str(entidad))
                return entidad
                
        except KeyError as e:
            pytest.skip(
                f"Obtener Entidad: no se encontro la Key a la cual se hace referencia {elemento} en el archivo JSON de la pagina {page}. Error: {e}"
            )         
            
    # Método genérico para obtener un elemento a partir del archivo JSON
    def Get_Element(self,page, elemento):
        GetEntity = Functions.Get_Entity(self, page, elemento)
        if GetEntity is None:
            print(u'No se encontro el valor de la entidad buscada en el archivo .Json')
        else:
            try:
                elemento = Functions.Find_Element_On_DOM(self, GetEntity["GetFieldBy"].upper(), GetEntity["ValueToFind"]) 
                print(u'Obtener Elemento: se encontro el elemento: ' + GetEntity["GetFieldBy"] + ' con el valor: ' + GetEntity["ValueToFind"])
                return elemento
            except NoSuchElementException:
                print(u'Obtener Elemento: no se encontro el elemento ' + GetEntity["GetFieldBy"] + ' con el valor: ' + GetEntity["ValueToFind"])
                Functions.cerrar_driver_navegador(self)
            except TimeoutException:
                print(u'Obtener Elemento: no se encontro el elemento ' + GetEntity["GetFieldBy"] + ' con el valor: ' + GetEntity["ValueToFind"])
                Functions.cerrar_driver_navegador(self)          
    
    # Método para obtener el texto de un elemento a partir de la entidad del archivo JSON
    def Get_Text(self,page, elemento):
        GetEntity = Functions.Get_Element(self, page, elemento)
        try:
            print(u'Obtener Texto: Texto en el elemento ' + str(elemento))
            return GetEntity.text
        except NoSuchElementException:
            print(u'Obtener Texto: No se logro obtener el texto del elemento ' + str(elemento))
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Obtener Texto: No se logro obtener el texto del elemento ' + str(elemento))
            Functions.cerrar_driver_navegador()

    # Método para hacer click en un elemento a partir de la entidad del archivo JSON 
    def Click_Element(self,page, elemento):
        GetEntity = Functions.Get_Element(self, page, elemento)
        try:
            print(u'Se realizo click en el elemento ' + str(elemento))
            return GetEntity.click()
        except NoSuchElementException:
            print(u'Elemento click: no se encontro el elemento ' + str(elemento) + u' para hacer click')
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Elemento click, no se encontro el elemento ' + str(elemento) + u' para hacer click')
            Functions.cerrar_driver_navegador(self)
    
    # Método para enviar texto a un elemento a partir de la entidad del archivo JSON
    def SendKeys(self,page,elemento,texto):
        GetEntity = Functions.Get_Element(self, page, elemento)
        try:
            print(f'Escribir texto: se escribio el texto {texto} en el elemento {elemento}')
            return GetEntity.send_keys(texto)
        except NoSuchElementException:
            print(u'Escribir Texto: No se encontro el elemento ' + str(elemento) + u' para escrbir el valor: {texto}')
            Functions.cerrar_driver_navegador()
            
        except TimeoutException:
            print(u'Escribir Texto: No se encontro el elemento ' + str(elemento) + u' para escribir el valor: {texto}')
            Functions.cerrar_driver_navegador()
    
    # Método para enviar teclas específicas a un elemento a partir de la entidad del archivo JSON
    def Send_Keys_Specific(self,page,elemento,key):
        try:
            if key.lower()=='enter':
                Functions.Get_Element(self,page,elemento).send_keys(Keys.ENTER)
                print(u'Se presiono la tecla ' + key + ' en el elemento indicado: ' + str(elemento))
            if key.lower()=='tab':
                Functions.Get_Element(self,page,elemento).send_keys(Keys.TAB)
                print(u'Se presiono la tecla ' + key + ' en el elemento indicado: ' + str(elemento))
            if key.lower()=='space':
                Functions.Get_Element(self,page,elemento).send_keys(Keys.SPACE)   
                print(u'Se presiono la tecla ' + key + ' en el elemento indicado: ' + str(elemento))
                
        except TimeoutException:  
            print(u'No se logro realizar la acción con la tecla indicada ' + key + " en el elemento indicado: " + str(elemento))
            Functions.cerrar_driver_navegador(self)   

    # Método para limpiar el texto de un elemento a partir de la entidad del archivo JSON
    def Clear_Element(self,page, elemento):  
        GetEntity = Functions.Get_Element(self, page, elemento)
        try:
            print(u'Limpiar Elemento: Se limpio el elemento ' + str(elemento))
            return GetEntity.clear()
        except NoSuchElementException:
            print(u'Limpiar Elemento: no se logro limpiar el elemento ' + str(elemento))
            Functions.cerrar_driver_navegador(self)   
        except TimeoutException:
            print(u'Limpiar Elemento: no se logro limpiar el elemento ' + str(elemento))
            Functions.cerrar_driver_navegador(self)

    # Método para obtener un elemento select a partir de la entidad del archivo JSON
    def Get_Element_Select(self,page, elemento):
        GetEntity = Functions.Get_Entity(self, page, elemento)
        if GetEntity is None:
            print(u'No se encontro el valor de la entidad buscada en el archivo .Json')
        else:
            try:
                select = Select(self.driver.find_element(Functions.BY[GetEntity["GetFieldBy"].upper()],GetEntity["ValueToFind"]))
                print(u"get elements: " + GetEntity["ValueToFind"]) 
                return select
            except NoSuchElementException:
                print(u"Select_Element: No presente " + GetEntity["ValueToFind"])
                Functions.cerrar_driver_navegador(self) 
            except TimeoutException:
                print(u"Select_Element: No presente " + GetEntity["ValueToFind"])
                Functions.cerrar_driver_navegador(self)

    # Método para seleccionar un elemento de un select por su texto visible a partir de la entidad del archivo JSON  
    def obtener_elemento_select_texto(self,page, elemento, texto):
        select = Functions.Get_Element_Select(self, page, elemento)
        select.select_by_visible_text(texto)
    
    # Método para esperar explícitamente a que un elemento sea visible y clickeable a partir de la entidad del archivo JSON          
    def Explicit_Wait_Element(self, page, elemento, tiempo_espera):  
        GetEntity = Functions.Get_Entity(self, page, elemento)
        if GetEntity is None:
            print(u'No se encontro el valor de la entidad buscada en el archivo .Json')
        else:
            try:
                wait =WebDriverWait(self.driver,tiempo_espera)
                wait.until(EC.visibility_of_element_located((Functions.BY[GetEntity["GetFieldBy"].upper()],GetEntity["ValueToFind"])))
                wait.until(EC.element_to_be_clickable((Functions.BY[GetEntity["GetFieldBy"].upper()],GetEntity["ValueToFind"])))   
                print(u'Espera explicita: se visualizo el elemento ' + str(page) + ' con el valor ' + GetEntity["ValueToFind"])
                return True
            except NoSuchElementException:
                print(u'Esperar explicita: no se encontro o no se visualizo el elemento luego de la espera ' + str(page) + ' con el valor ' + GetEntity["ValueToFind"])
                Functions.cerrar_driver_navegador()
            except TimeoutException:
                print(u'Esperar explicita: no se encontro o no se visualizo el elemento luego de la espera ' + str(page) + ' con el valor ' + GetEntity["ValueToFind"])
                Functions.cerrar_driver_navegador()

    # Método para realizar scroll hasta un elemento a partir de la entidad del archivo JSON
    def Scroll_Element_JS(self, page, elemento):
        GetEntity = Functions.Get_Element(self, page, elemento)
        try: 
            self.driver.execute_script("arguments[0].scrollIntoView();", GetEntity)
            print(u'JS Scroll: Se realizo scroll_to hasta el elemento ' + str(page) + ' con el valor ' + elemento)
            return True
        except TimeoutException:
            print(u'JS Scroll: No se logro realizar hacia el elemento ' + str(page) + ' con el valor ' + elemento)
            Functions.cerrar_driver_navegador(self)

    # Método para hacer doble click en un elemento a partir de la entidad del archivo JSON
    def Double_Click(self, page, elemento):
        action =ActionChains(self.driver)
        element = Functions.Get_Element(self, page, elemento)
        try:
            action.double_click(element).perform()
            print(f'Double Click: se realizo doble click en el elemento ' + str(page) + ' con el valor ' + str(elemento))

        except NoSuchElementException:
            print(u'Double Click: no se logro realizar doble click en el elemento ' + str(page) + ' con el valor ' + str(elemento))
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Double Click: no se logro realizar doble click en el elemento ' + str(page) + ' con el valor ' + str(elemento))
            Functions.cerrar_driver_navegador(self)

    # Método para hacer click derecho en un elemento a partir de la entidad del archivo JSON    
    def Click_Derecho(self,page, elemento):
        action =ActionChains(self.driver)
        element = Functions.Get_Element(self, page, elemento)
        try:
            action.context_click(element).perform()
            print(f'Click Derecho: se realizo click derecho en el elemento ' + str(page) + ' con el valor ' + str(elemento))
        except NoSuchElementException:
            print(u'Click Derecho: no se logro realizar click derecho en el elemento ' + str(page) + ' con el valor ' + str(elemento))
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Click Derecho: no se logro realizar click derecho en el elemento ' + str(page) + ' con el valor ' + str(elemento))
            Functions.cerrar_driver_navegador(self)

    # Método para mover el mouse a un elemento a partir de la entidad del archivo JSON
    def Move_Element(self,page, elemento):
        action = ActionChains(self.driver)
        element = Functions.Get_Element(self, page, elemento)
        try:
            action.move_to_element(element).perform()
            print(f'Mover Mouse: se movio el mouse al elemento ' + str(page) + ' con el valor ' + str(elemento))
        except NoSuchElementException:
            print(u'Mover_mouse: no se encontro el elemento ' + str(page) + ' con el valor ' + str(elemento))
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Mover_mouse: no se encontro el elemento ' + str(page) + ' con el valor ' + str(elemento))
            Functions.cerrar_driver_navegador(self)  

    #Espera informal
    def esperar_elemento(self,tiempo_espera = Inicializar.Tiempo_Espera):
        print("Inicia Espera: " +str(tiempo_espera))
        try:
            totalWait = 0
            while(totalWait < tiempo_espera):
                time.sleep(1)
                totalWait = totalWait + 1
                print("Tiempo total actual de espera: " + str(totalWait))
        finally:
            print("Espera: Carga Finalizada")

    #Crear libro excel
    def crear_libro_excel(self,celda1=None, celda2=None, celda3=None,  celda4=None, celda5=None):
        #crear el libro
        wb = openpyxl.Workbook()
        
        #seleccionar e imprimir la hoja activa
        hoja = wb.active
        print(f'Hoja activa: {hoja.title}')
        
        #renombrar e imprimir la hoja activa
        hoja.title = "DataTest"
        print(f'Hoja activa: {hoja.title}') 
        
        hoja =wb["DataTest"]  
        if (celda1 != None)and (celda2 != None) and (celda3 != None):
            hoja[celda1] = "Valor Tomado"
            hoja[celda2] = "Valor Escrito"
            hoja[celda3] = "Resultado"
        
        if (celda4 != None)and (celda5 != None):
            hoja1 = wb.create_sheet("Bitacora Pruebas")
            hoja1[celda4] = "Prueba"
            hoja1[celda5] = "Resultado" 
            
        RutaArchivo = Inicializar.Excel_Crear + "\Pruebas.xlsx"     
        print(wb.sheetnames)
        print(u'Se creó el archivo en la ruta: ' + RutaArchivo)
        wb.save(RutaArchivo)     
    
    #Leer celda de excel   
    def leer_celda(self,celda, hoja):
        wb = openpyxl.load_workbook(Inicializar.Excel_Leer_Escribir)
        sheet = wb[hoja]
        valor = sheet[celda].value
        print(u'--------------------------------------------------')
        print(u'El libro de excel utilizado es: ' + Inicializar.Excel_Leer_Escribir)
        print(f'El valor de la celda es: {valor}')
        print(u'--------------------------------------------------')  
        return valor
    
    #Escribir en celda de excel
    def escribir_celda(self,celda,hoja,valor):
        wb = openpyxl.load_workbook(Inicializar.Excel_Leer_Escribir)
        hoja = wb[hoja]
        hoja[celda]=valor
        wb.save(Inicializar.Excel_Leer_Escribir)
        print(u'--------------------------------------------------')
        print(u'El libro de excel utilizado es: ' + Inicializar.Excel_Leer_Escribir)
        print(u'Se escribio en la celda: '+ celda + ' el valor: ' + valor)
        print(u'--------------------------------------------------')  
    
    def WebdriverWait(self,time):
        WebDriverWait(self.driver,time)
    
    #Obtener fecha actual
    def obtener_fecha_actual(self):
        self.fecha = time.strftime(Inicializar.DateFormat)#Formato Fecha
        return self.fecha
    
    #Obtener hora actual
    def obtener_hora_actual(self):
        self.hora = time.strftime(Inicializar.HourFormat)#Formato 24Hrs
        return self.hora
    
    #Crear ruta para capturas de pantallas
    def crear_path(self):
        fecha = Functions.obtener_fecha_actual(self)
        GeneralPath = Inicializar.Path_Evidencias
        print(f'Ruta General de las Capturas: {GeneralPath}')
        DriverTest = self.Nav_utilizado_capturas            
        TestCase =self.__class__.__name__

        HoraActual = Functions.obtener_hora_actual(self)
        
        if   ((Inicializar.TestCase_x_Context =="S") and (GeneralPath != "")):
            path = f"{GeneralPath}\{fecha}\Pruebas\{TestCase}\{DriverTest}\{HoraActual}"
            print(f"Ruta Contruida para guardar las capturas: {path}")
        elif ((Inicializar.TestCase_x_Context == "N") and (GeneralPath != "")):
            path =f"{GeneralPath}\{fecha}\{TestCase}\{DriverTest}\{HoraActual}"
            print(f"Ruta Contruida para guardar las capturas: {path}")
        elif (((Inicializar.TestCase_x_Context == "N") or (Inicializar.TestCase_x_Context == "S")) and (GeneralPath == "")):
            path = f'{Inicializar.BaseDir}\Capturas\{fecha}\{TestCase}\{DriverTest}\{HoraActual}'
            print(f'No se encuentra establecida la ruta para guardar la captura de pantalla, se guardara en la carpeta raiz del framework de pruebas.\nEn: {path}')
        elif (((Inicializar.TestCase_x_Context !="S") or (Inicializar.TestCase_x_Context !="N") or (Inicializar.TestCase_x_Context == "")) and (GeneralPath == "")): 
            path = ""
            print(f'No se logro crear el path para guardar la captura de pantalla. Variables de "TestCase_x_Context" y "GeneralPath" no se han configurado correctamente: Tienen el valor: {GeneralPath} y {Inicializar.TestCase_x_Context}')

        if (path != ""):
            if not os.path.exists(path):
                os.makedirs(path)
            
            return path
        else:
            return Inicializar.Warning_Evidencias
        
    def crear_path_evidencias_video(self):
        fecha = Functions.obtener_fecha_actual(self)
        GeneralPath = Inicializar.Path_Videos
        print(f'Ruta General de las Capturas: {GeneralPath}')
        DriverTest = self.Nav_utilizado_capturas
        TestCase =self.__class__.__name__
        HoraActual = Functions.obtener_hora_actual(self)
        if (Inicializar.Path_Videos != ""):
            path = f"{GeneralPath}\{fecha}\{TestCase}\{DriverTest}\{HoraActual}"
            print(f"Ruta Contruida para guardar los videos: {path}")
        elif (Inicializar.Path_Videos == ""):
            path = f'{Inicializar.BaseDir}\Videos\{fecha}\{TestCase}\{DriverTest}\{HoraActual}'
            print(f'No se encuentra establecida la ruta para guardar los videos, se guardara en la carpeta raiz del framework de pruebas.\nEn: {path}')

        if (path != ""):
            if not os.path.exists(path):
                os.makedirs(path)
                return path
            else:
                return Inicializar.Warning_Evidencias

    #Realizar captura de pantalla
    def capturar_pantalla(self):
        Path=Functions.crear_path(self)
        TestCase =self.__class__.__name__
        
        if Path != Inicializar.Warning_Capturas:
            img = f'{Path}\{TestCase}\
            ('+Functions.obtener_fecha_actual(self)+' - '+ Functions.obtener_hora_actual(self)+')'+'.png'
            
            print(f'Se realizo captura de pantalla de la prueba: {img}')
            return self.driver.get_screenshot_as_file(img)
        else:
            print("Warning: No se logro generar la captura de pantalla. No se encuentra configurada el Path y variable contexto.")
    
    #Realizar captura de pantalla en reporte Allure
    def captura_pantalla_allure(self,Descripcion):
        allure.attach(self.driver.get_screenshot_as_png(),Descripcion,allure.attachment_type.PNG)
    
    #Realizar conexion a BD     
    def pyodbc_conexionBD(self,Env):
            
            try:
                if Env == 'DEV':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_Dev)
                elif Env == 'QA':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_QA)
                elif Env == 'UAT':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_UAT)
                elif Env == 'PROD':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_Prod)
                else:
                    print('No se logro establecer la cadena de conexion con la base de datos.')
                    
                self.cursor = conn.cursor()
                print("Conexion Exitosa")
                return self.cursor
            except (pyodbc.OperationalError) as Error:
                self.cursor = None
                print("Conexion Fallida")
                pytest.skip("Error en la conexion a la BD: ", str(Error))
    
    #Realizar consulta a BD         
    def pyodbc_ConsultaBD(self,Env,consulta_query):
        self.cursor = Functions.pyodbc_conexionBD(self,Env)
        if self.cursor is not None:
            try:
                self.cursor.execute(consulta_query)
                self.Result = self.cursor.fetchall()
                for row in self.Result:
                    return row[0]
                print(row[0])
                
            except (pyodbc.Error) as Error:
                print('Error en la consulta: ', Error)
                
            finally:
                if(self.cursor):
                    self.cursor.close()
                    print('pyodbc: Se cerro la conexion con la BD')                 
    
    #Abrir nuevo tab en instancia del navegador
    def abrir_nuevo_tab(self,Name_Tab,Page_Tab=Inicializar.Page_Tab): 
        self.driver.execute_script(f'''window.open("{Page_Tab}","{Name_Tab}");''')
        #self.driver.switch_to.new_window("tab")
        print(f'''window.open("{Page_Tab}","{Name_Tab}");''')
    
    #Abrir nuevo tab en nueva ventana del navegador
    def abrir_nuevo_tab_en_nueva_ventana(self,URL,Name_Tab,Page_Tab = Inicializar.Page_Tab):
        self.driver.execute_script(f'''window.open("{URL}","{Page_Tab}","{Name_Tab}");''')
        print(f'''window.open("{URL}","{Page_Tab}","{Name_Tab}");''')
    
    #Espera explicita
    def espera_explicita(self,driver,time, page_state):
        WebDriverWait(driver,time).until(lambda driver: page_state == 'complete')
        assert page_state == 'complete','No se completo la carga del sitio'
    
    #Moverse o intercambiar entre tabs abiertos en el navegador
    def intercambio_tab(self,ventana):
        self.driver.switch_to.window(self.driver.window_handles[ventana])
    
    #Cargar Archivo
    def cargar_archivo(self,entidad,txt_ruta_archivo = Inicializar.Archivo_Cargar):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        if Get_Entidad is None:
            return print(u'No se encontro el valor de la entidad requerida en el doc. Json')
        else:
            try:
                if self.json_GetFieldBy.lower()=='id':
                    Functions.escribir_texto(self, *entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='name':
                    Functions.escribir_texto(self, *entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='xpath':
                    Functions.escribir_texto(self, *entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                    
            except NoSuchElementException:
                print(u'No se pudo cargar el archivo en el elemento: ' + entidad + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
            except TimeoutException:
                print(u'No se pudo cargar el archivo en el elemento: ' + entidad + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)   
    
    #Assert
    def Assert_Equal(self,*elemento,texto_esperado,msj):
        return self.assertEqual(Functions.obtener_Texto(self, *elemento),texto_esperado, msj)
    def Assert_True(self,*elemento,texto,msj):
        return self.assertTrue(Functions.obtener_Texto(self, *elemento)==texto, msj)
    def AssertFalse(self,*elemento,texto,msj):
        return self.assertFalse(Functions.obtener_Texto(self, *elemento)==texto, msj)
    def Assert_True_IsDisplayer(self,*elemento, msj):
        elemento = Functions.obtener_elemento(self, *elemento)
        return self.assertTrue(elemento.is_displayed()==True,msj)
    def Assert_In_Elemento(self,texto_contenido, texto_ingresado_alert):
        return self.assertIn(texto_contenido, f'{texto_ingresado_alert}')

    #Arrastrar y Soltar: Drag & Drop
    def Arrastrar_y_Soltar(self,elemento_drag, elemento_drop):
        action = ActionChains(self.driver)
        element_drag = Functions.obtener_elemento(self, elemento_drag) 
        elemento_drop = Functions.obtener_elemento(self,elemento_drop) 
        
        try:
            #Perform drap and drop
            action.drag_and_drop(element_drag, elemento_drop).perform()  
        except NoSuchElementException:
            print(u'Drap_and_Drop: no se logro realizar la accion entre los elementos ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Drap_and_Drop: no se logro realizar la accion entre los elementos ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
    
    #Mover Mouse entre elementos en aplicativo web       
    def Mover_Mouse(self,elemento1, elemento2):
        action =ActionChains(self.driver)
        element1 = Functions.obtener_elemento(self, elemento1)
        element2 = Functions.obtener_elemento(self, elemento2)
        
        try:
            #Mover Cursor
            action.move_to_element(element1).move_to_element(element2).perform()
        except NoSuchElementException:
            print(u'Mover Cursor: no se logro realizar la accion ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Mover Cursor: no se logro realizar la accion ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
    
    #Descargar Archivo
    def download_file(self,*elemento):
        Functions.esperar_elemento(self)
        Functions.click_en_elemento(self, *elemento)
        Functions.esperar_elemento(self)
        contenido = os.listdir(Inicializar.Ruta_Descarga)
        print(contenido)
        Functions.Assert_True_IsTrueDownload(self, contenido, f'No se descargo el archivo {Inicializar.Archivo_Descargado} en la ruta {Inicializar.Ruta_Descarga}')
    
    #Validar la descarga del archivo     
    def Assert_True_IsTrueDownload(self,contenido, msj):
        return self.assertTrue(Inicializar.Archivo_Descargado in contenido,msj)
        Functions.esperar_elemento(self)           
    
    #Escribir en un archivo bitacora
    def write_file_txt(self,texto,archivo):
        bitacora_pruebas = open(archivo, 'a')
        bitacora_pruebas.write(f'\nPrueba: {Functions.obtener_fecha_actual(self)} - {Functions.obtener_hora_actual(self)} - {texto} \n')
        bitacora_pruebas.write(f'--------------------------------------------------------------------')
        bitacora_pruebas.close()
    
    #Obtener cookie x nombre
    def obtener_cookie_nombre(self,nombre_cookie):
        print('--------------- Cookie x Nombre ----------------')
        cookie = self.driver.get_cookie(nombre_cookie) 
        print(cookie)
    
    #Obtener todas las cookies  
    def obtener_todas_las_cookies(self):
        print('--------------- Todas las cookies ----------------')
        cookie = self.driver.get_cookies() 
        print(cookie)      
    
    #Eliminar cookies x nombre
    def eliminar_cookie_x_nombre(self,nombre_cookie):
        print('-------------- Eliminar Cookie x Nombre ------------')
        self.driver.delete_cookie(nombre_cookie)
    
    #Eliminar todas las cookies     
    def eliminar_todas_las_cookies(self):
        print('-------------- Eliminar todas las cookies ---------------')
        self.driver.delete_all_cookies()
     
    #Realizar captura y cortar imagen   
    def captura_de_imagen_cortada(self, elemento_imagen, Ruta_Img_Cortada = Inicializar.Imagenes_Cortadas):
        imagen_encontrada = elemento_imagen.location
        size = elemento_imagen.size
        
        print(size)
        print(imagen_encontrada)
        
        imagen_guardada = self.driver.get_screenshot_as_png()
        imagen_cortada = Image.open(BytesIO(imagen_guardada))
        left = imagen_encontrada['x']
        top = imagen_encontrada['y']
        right = imagen_encontrada['x'] + size['width']
        bottom = imagen_encontrada['y'] +size['height']
        
        print(left,top,right,bottom)
        
        imagen_cortada = imagen_cortada.crop((left,top,right,bottom))
        imagen_cortada.save(f'{Ruta_Img_Cortada}\imagen_cortada.png')
    
    #Seleccionar fechas DTimePicker Dinamico
    def Selects_Fechas_DTPickerDinamico(self,AvanzarMes,Meses_Avanzar,FechaIda,FechaVuelta,DiaIda,DiaVuelta):
      #Contador para avanzar en el calendario
      avanzar = 0  
      
      #Click en control fecha ida
      Functions.click_en_elemento(self, FechaIda)
      Functions.esperar_elemento(self)
      
      #Avanzar en los meses del calendario     
      while (avanzar < Meses_Avanzar):
          Functions.click_en_elemento(self,AvanzarMes)
          avanzar = avanzar+1
    
      Functions.click_en_elemento(self, DiaIda)
      Functions.esperar_elemento(self)
      avanzar = 0
      
      #Avanzar en los meses del calendario  
      while (avanzar < Meses_Avanzar):
          Functions.click_en_elemento(self,AvanzarMes)
          avanzar = avanzar+1
          
      #Click en control fecha vuelta
      Functions.click_en_elemento(self, FechaVuelta)
      Functions.esperar_elemento(self)
      
      #Seleccionar dia vuelta
      Functions.click_en_elemento(self, DiaVuelta)
      Functions.esperar_elemento(self)
      
    #Seleccionar fechas DTimePicker Dinamico
    def Select_Fechas_DTPickerDinamicoUnico(self,FechaIda,Dia): 
      
        #Click en control fecha ida
        Functions.click_en_elemento(self, FechaIda)
        Functions.esperar_elemento(self)
        
        Functions.click_en_elemento(self, Dia)
        Functions.esperar_elemento(self)
     
    #Metodos para grabar videos (screen record) en las pruebas.  
    def configurar_entorno_grabacion(self): 
        self.recording = True
        self.sct = MSS()
        self.monitor = self.sct.monitors[1]
        self.fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        
        self.video_path = Functions.crear_path_evidencias_video(self)
        self.nombre_video = f"{self.video_path}\{Functions.obtener_fecha_actual(self)}_{Functions.obtener_hora_actual(self)}_recording.mp4"

        if self.video_path != Inicializar.Warning_Evidencias:
            self.out = cv2.VideoWriter(
                self.nombre_video,
                self.fourcc,
                20.0,
                (self.monitor["width"], self.monitor["height"])
                )

            self.recording_thread = threading.Thread(
                target= self.record_screen,
                daemon=True
            )       
            self.recording_thread.start()
        else:
            print(Inicializar.Warning_Evidencias)

    def record_screen(self):
        while self.recording:

            img = np.array(
                self.sct.grab(self.monitor)
            )

            frame = cv2.cvtColor(
                img,
                cv2.COLOR_BGRA2BGR
            )
            self.out.write(frame)
            time.sleep(0.01)
            
    def detener_grabacion(self):
        self.recording = False
        self.recording_thread.join(timeout=2)
        self.out.release()

    def alert_navegadores(self,tipo_alert,texto_esperado="", msj="",texto_contenido ="", elemento="",texto_ingresado=""):
        self.alert = Alert(self.driver)
        self.text_alert = self.alert.text
        #0= alert OK por defecto, 1= alert espera de 5 segundos, 2=  ok y cancel, 3=  ok y cancel, 4= texto ingresado    
        #Alert con boton OK
        if (tipo_alert == 0):
            print('Mensaje de Alert: ' + self.text_alert)
            self.alert.accept()
        #Alert con espera de tiempo
        elif(tipo_alert == 1):
            print('Mensaje de Alert: ' + self.text_alert)
            self.alert.accept()
        #Alert con botones OK y Cancel
        elif(tipo_alert == 2):
            print('Mensaje de Alert: ' + self.text_alert)
            self.alert.accept()
            Functions.Assert_Equal(self,elemento, texto_esperado, msj)
        #Alert con botones OK y Cancel
        elif(tipo_alert == 3):
            print('Mensaje de Alert: ' + self.text_alert)
            self.alert.dismiss()
            Functions.Assert_Equal(self,elemento, texto_esperado, msj)
        #Alert con ingreso de texto
        elif(tipo_alert == 4):
            print('Mensaje de alert: ' + self.text_alert)
            self.alert.send_keys(texto_ingresado)
            self.alert.accept()
            Functions.Assert_In_Elemento(self,texto_contenido, Functions.obtener_Texto(self, elemento))    